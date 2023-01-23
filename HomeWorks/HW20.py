# Прочитать сохранённый csv-файл из задания №19 и сохранить данные в excel-файл,
# кроме возраста – столбец с этими данными не нужен.

import csv
import openpyxl
import random

with open('HW19.csv', mode="r") as file_csv:
    read_file = csv.reader(file_csv, delimiter=",")
    count = 0
    fields = []
    mob_codes = ["093", "067", "073"]
    for index, row in enumerate(read_file):
        del row[2]
        if index > 0:
            row.insert(0, "Person-{0}".format(index))
            if index % 3 != 0:
                mob_code = mob_codes[random.randint(0, 2)]
                row[3] = "+38" + mob_code + str(random.randint(0, 999999)).zfill(6)
        else:
            row.insert(0, "Name of fields")
        fields.append(row)
        count += 1
    print(fields)
    print(count)

wb = openpyxl.Workbook()
wb.create_sheet(title="First", index=0)
worksheet = wb['First']
for col_index, col in enumerate(fields):
    for row_index, value in enumerate(col):
        cell = worksheet.cell(column=col_index + 1, row=row_index + 1)
        cell.value = value
wb.save('HW20.xlsx')
