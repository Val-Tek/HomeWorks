import openpyxl
from datetime import datetime
from time import sleep
import os, fnmatch


class Person(object):
    def __init__(self, first_name, date_of_birth, gender, middle_name=None, last_name=None, date_of_death=None):
        self.first_name = first_name
        self.middle_name = middle_name
        self.last_name = last_name
        self.date_of_birth = date_of_birth
        self.gender = gender
        self.date_of_death = date_of_death

    def __str__(self):
        if self.date_of_death is None:
            return f"{self.full_name()}, {self.gender}, {self.age()} years old, was born {self.date_of_birth} ."
        else:
            return f"{self.full_name()}, {self.gender}, {self.age()} years old, was born {self.date_of_birth}, " \
                   f"dead {self.date_of_death}."

    def full_name(self):
        full_name = f"{self.first_name}  {self.middle_name}  {self.last_name}"
        return full_name

    def age(self):
        y = self.date_of_birth
        if self.date_of_death is None:
            x = datetime.now()
        else:
            x = self.date_of_death

        if x.month - y.month > 0:
            result = x.year - y.year
        else:
            if x.day - y.day >= 0:
                result = x.year - y.year
            else:
                result = (x.year - y.year) - 1
        return result


def main():
    # file selection loop
    print("Welcome to the data saver programme!")
    sleep(1)
    persondata = []
    file_name = ""
    while True:
        s = input("Enter 'N' to create a new file) or (enter 'L' to load existing file) or (enter 'Q' to quit) : ")
        if s in ["Q", "q"]:
            print("Thank you for using my programme!")
            return
        if s in ["N", "n"]:
            file_name = input("Enter a name for new file: ")
            break
        if s in ["L", "l"]:
            file_name = input(f'Enter the file_name from project {files_names()}: ')
            persondata = load_file(file_name)
            print(f"{len(persondata)} records loaded.")
            break
        else:
            print("Command not recognized :(")
            # continue file selection loop

    # file view-editing loop
    while True:
        print("*" * 50)
        sleep(1)
        print("You have available commands: ")
        sleep(1)
        print(" to view content - 'V' \n to enter new data -'E'\n to search in data - 'S' \n to quit - 'Q'")
        sleep(2)
        cmd = input("Pleas, enter one of the commands- V,E,S,Q: ")
        if cmd in ["V", "v"]:
            view_data(persondata)
        elif cmd in ["E", "e"]:
            new_person = enter_Record()
            persondata.append(new_person)
        elif cmd in ["S", "s"]:
            result = search(persondata)
            for p in result:
                print(p)
        elif cmd in ["Q", "q"]:
            q_cmd = input("do you want to save the changes? enter: Y (yes) or N (no): ")
            if q_cmd in ["Y", "y"]:
                save_file(file_name, persondata)
                print(f"File {file_name} saved  successfully :)")
                break
            if q_cmd in ["N", "n"]:
                print("Thank you for using my programme!")
                break
        else:
            print("Command not recognized :(")
            # continue view-editing  loop


def files_names():
    files_names = []
    listOfFiles = os.listdir('.')
    pattern = "*.xlsx"
    for file in listOfFiles:
        if fnmatch.fnmatch(file, pattern):
            files_names.append(file)
    return files_names


def load_file(file_name):
    wb = openpyxl.load_workbook(file_name)
    worksheet = wb.active
    persondata = []
    for i in range(2, worksheet.max_row + 1):
        middlename = worksheet.cell(row=i, column=2).value
        if middlename is None:
            middlename = ""
        lastname = worksheet.cell(row=i, column=3).value
        if lastname is None:
            lastname = ""
        person = Person(first_name=worksheet.cell(row=i, column=1).value,
                        middle_name=middlename,
                        last_name=lastname,
                        date_of_birth=worksheet.cell(row=i, column=4).value,
                        date_of_death=worksheet.cell(row=i, column=5).value,
                        gender=worksheet.cell(row=i, column=6).value)
        persondata.append(person)
    return persondata


def save_file(file_name, persondata):
    wb = openpyxl.Workbook()
    wb.create_sheet(title="First", index=0)
    worksheet = wb.active
    worksheet.cell(row=1, column=1).value = "first_name"
    worksheet.cell(row=1, column=2).value = "middle_name"
    worksheet.cell(row=1, column=3).value = "last_name"
    worksheet.cell(row=1, column=4).value = "date_of_birth"
    worksheet.cell(row=1, column=5).value = "date_of_death"
    worksheet.cell(row=1, column=6).value = "gender"
    i = 2
    for person in persondata:
        worksheet.cell(row=i, column=1).value = person.first_name
        worksheet.cell(row=i, column=2).value = person.middle_name
        worksheet.cell(row=i, column=3).value = person.last_name
        worksheet.cell(row=i, column=4).value = person.date_of_birth
        worksheet.cell(row=i, column=5).value = person.date_of_death
        worksheet.cell(row=i, column=6).value = person.gender
        i += 1
    wb.save(file_name)


def view_data(data):
    print(f"We have {len(data)} Person records;")
    for i in data:
        print(i)


def enter_Record():
    first_name = input("first_name: ")
    middle_name = input("middle_name: ")
    last_name = input("last_name: ")
    gender = input("gender: ")
    while True:
        strinput = input("Enter date of birth dd.mm.yyyy: ")
        try:
            date_of_birth = datetime.strptime(strinput, "%d.%m.%Y")
            break
        except ValueError:
            print("Format not valid try again dd.mm.yyyy ")
    while True:
        strinput = input("Enter date of death dd.mm.yyyy: ")
        if strinput == "":
            date_of_death = None
            break
        try:
            date_of_death = datetime.strptime(strinput, "%d.%m.%Y")
            break
        except ValueError:
            print("Format not valid try again dd.mm.yyyy ")
    p = Person(first_name=first_name, date_of_birth=date_of_birth, gender=gender, middle_name=middle_name,
               last_name=last_name, date_of_death=date_of_death)
    return p


def search(data):
    search_data = input("Enter the request for searching by names in data: ")
    result = []
    for person in data:
        if search_data.lower() in person.full_name().lower():
            result.append(person)
        else:
            result = ["Request not found."]
    return result


main()
